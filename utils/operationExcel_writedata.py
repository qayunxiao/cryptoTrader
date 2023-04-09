# -*- coding: utf-8 -*-
# @Time    : 2019/9/18 14:41
# @Author  : alvin
# @File    : Data_Excel_Action.py
# @Software: PyCharm

import os
import time
import xlrd
from xlutils.copy import copy
from utils.public import *
from utils.excel_data_cpadataimp import *
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook


class Data_Excel_Action( object ):

    def getExcel(self,fileName):
        print("getExcel  data_dir_file",data_dir_file( fileName=fileName ))

    def get_rows(self):
        '''获取excel的行数'''
        return self.getExcel().nrows

    def get_row_cel(self, row, col):
        '''获取单元格的内容'''
        return self.getExcel().cell_value( row, col )

    def get_date(self, row):
        return self.get_row_cel( row, get_date() )

    def get_ad_type(self, row):
        return self.get_row_cel( row, get_ad_type() )

    def get_ad_name(self, row):
        return self.get_row_cel( row, get_ad_name() )

    def get_ad_package(self, row):
        return self.get_row_cel( row, get_ad_package() )

    def get_ad_channel(self, row):
        return self.get_row_cel( row, get_ad_channel() )

    def get_ad_activenum(self, row):
        return self.get_row_cel( row, get_ad_activenum() )

    def get_ad_active_remark(self, row):
        return self.get_row_cel( row, get_ad_active_remark() )

    def write_date(self, row, content):
        '''广告名称写到文件中'''
        col = get_date()
        work = xlrd.open_workbook(
            data_dir_file( fileName="export_cpa_demo.xlsx" ) )
        old_content = copy( work )
        ws = old_content.get_sheet( 0 )
        ws.write( row, col, content )

    def write_ad_type(self, row, content):
        '''广告名称写到文件中'''
        col = get_ad_type()
        work = xlrd.open_workbook(
            data_dir_file( fileName="export_cpa_demo.xlsx" ) )
        old_content = copy( work )
        ws = old_content.get_sheet( 0 )
        ws.write( row, col, content )

    def write_ad_name(self, row, content):
        '''广告名称写到文件中'''
        col = get_ad_name()
        work = xlrd.open_workbook(
            data_dir_file( fileName="export_cpa_demo.xlsx" ) )
        old_content = copy( work )
        ws = old_content.get_sheet( 0 )
        ws.write( row, col, content )

    def write_ad_package(self, row, content):
        '''广告名称写到文件中'''
        col = get_ad_package()
        work = xlrd.open_workbook(
            data_dir_file( fileName="export_cpa_demo.xlsx" ) )
        old_content = copy( work )
        ws = old_content.get_sheet( 0 )
        ws.write( row, col, content )

    def write_ad_channel(self, row, content):
        '''广告名称写到文件中'''
        col = get_ad_channel()
        work = xlrd.open_workbook(
            data_dir_file( fileName="export_cpa_demo.xlsx" ) )
        old_content = copy( work )
        ws = old_content.get_sheet( 0 )
        ws.write( row, col, content )

    def awrite_ethhistorydata(self,content):
        dataPath = os.path.join( os.path.dirname( os.path.dirname( __file__ ) ),'data' ,'MainBi')
        dataPath_file = os.path.join( dataPath, "mainbi_ethexchange.txt" )
        with open(dataPath_file,"a") as file:
            file.write(str(content)+'\n')

    def write_tokendata(self,content,filename):
        dataPath = os.path.join( os.path.dirname( os.path.dirname( __file__ ) ),'data' ,'tokeninfo')
        dataPath_file = os.path.join( dataPath,filename)
        with open(dataPath_file,"a") as file:
            file.write(str(content)+'\n')

    def write_alldata(self,content):
        # 获取文件路径
        dataPath = os.path.join( os.path.dirname( os.path.dirname( __file__ ) ),'data' ,'MainBi')
        # 定义文件名称
        #  invalid mode ('wb') or filename: 'Excel2017-09-21_20:15:57.xlsx'   这种方式明明文件，会提示保存失败，无效的文件名。
        # nameTime = time.strftime('%Y-%m-%d_%H:%M:%S')
        nameTime = time.strftime( '%Y-%m-%d_%H-%M-%S' )
        excelName = 'main_BI_list' + nameTime + '.xlsx'
        dataPath_file = os.path.join( dataPath, excelName )
        wb = Workbook()
        ws = wb.active
        tableTitle = ['日期','排名', '全名', '币种', '市值','币价RMB','币价USDT','换手率','BTC兑换个数']
        # 维护表头
        #        if row < 1 or column < 1:
        #          raise ValueError("Row or column values must be at least 1")
        # 如上，openpyxl 的首行、首列 是 （1,1）而不是（0,0），如果坐标输入含有小于1的值，提示 ：Row or column values must be at least 1，即最小值为1.
        # ws['A2'] = datetime.datetime( 2019,9,18)
        # print( "number_format",ws['A2'].number_format )  # yyyy-mm-dd h:mm:ss
        # ws['B3'] = '12%'
        # print( ws['B3'].number_format )  # General
        #写表头
        for col in range( len( tableTitle ) ):
            # print("len( tableTitle )",len( tableTitle ))
            c = col + 1
            ws.cell( row=1, column=c ).value = tableTitle[col]

        for row in range( len( content ) ):
            # print("len( content )",len( content ))
            ws.append( content[row] )

        wb.save( filename=dataPath_file )
        wb.close()
        return dataPath_file

    def write_data(self,fileName,content,tableTitle,locl='MainBi'):
        dataPath = os.path.join( os.path.dirname( os.path.dirname( __file__ ) ),'data' ,locl)
        nameTime = time.strftime( '%Y-%m-%d_%H-%M-%S' )
        excelName = fileName + nameTime + '.xlsx'
        dataPath_file = os.path.join( dataPath, excelName )
        wb = Workbook()
        ws = wb.active
        # tableTitle = ['日期','排名', '全名', '币种', '市值','币价RMB','币价USDT','换手率','BTC兑换个数']
        for col in range( len( tableTitle ) ):
            # print("len( tableTitle )",len( tableTitle ))
            c = col + 1
            ws.cell( row=1, column=c ).value = tableTitle[col]
        for row in range( len( content ) ):
            # print("len( content )",len( content ))
            ws.append( content[row] )
        wb.save( filename=dataPath_file )
        wb.close()
        return dataPath_file

    def write_ad_channel(self, row, content):
        '''广告名称写到文件中'''
        col = get_ad_channel()
        work = xlrd.open_workbook(
            data_dir_file( fileName="export_cpa_demo.xlsx" ) )
        old_content = copy( work )
        ws = old_content.get_sheet( 0 )
        ws.write( row, col, content )

    def write_edit_alldata(self,content,raw_filename):
        # 获取文件路径
        dataPath = os.path.join( os.path.dirname( os.path.dirname( __file__ ) ),'data' ,'upload_files')
        # 定义文件名称
        #  invalid mode ('wb') or filename: 'Excel2017-09-21_20:15:57.xlsx'   这种方式明明文件，会提示保存失败，无效的文件名。
        # nameTime = time.strftime('%Y-%m-%d_%H:%M:%S')
        nameTime = time.strftime( '%Y-%m-%d_%H-%M-%S' )
        excelName = 'export_cpa_demo' + nameTime + '.xlsx'
        dataPath_file = os.path.join( dataPath, excelName )
        # wb = Workbook()
        # file_name = data_dir_file( fileName="export_cpa_demo.xlsx")
        wb = load_workbook( raw_filename )
        ws = wb.active
       # 如上，openpyxl 的首行、首列 是 （1,1）而不是（0,0），如果坐标输入含有小于1的值，提示 ：Row or column values must be at least 1，即最小值为1.
        # ws['A2'] = datetime.datetime( 2019,9,18)
        # print( "number_format",ws['A2'].number_format )  # yyyy-mm-dd h:mm:ss
        # ws['B3'] = '12%'
        # print( ws['B3'].number_format )  # General
        for row in range( len( content ) ):
            # print("len( content )",len( content ))
            ws.append( content[row] )

        wb.save( filename=dataPath_file )
        wb.close()
        return dataPath_file

    def set_data_dateformat(self,file_name,rowsnum):
        '''
        http://blog.chinaunix.net/uid-23504396-id-4467013.html
        :param file_name: 需要操作的文件
        :param rowsnum: 需要修改日期的行总数
        :return: None
        '''
        wb = load_workbook( file_name )
        # # a_sheet = wb.get_sheet_by_name('Sheet')
        # # print("a_sheet title",a_sheet.title)
        sheet = wb.active
        rowsnum = rowsnum + 2
        for i in range (rowsnum):
            # print("i",i)
            if i >= 2:
                rowvalue = 'A'+str(i)
                date_col = sheet[rowvalue]
                date_col.number_format = 'yyyy/m/d'
            else:
                continue
        wb.save( file_name )
        print("set_data_dateformat is ok!")

if __name__ == "__main__":
    a = Data_Excel_Action()
    a.getExcel(r"D:\UserData\git\baice\qingniu\data\upload_files\export_cpa_demo2019-09-19_10-04-54.xlsx")
